# Підключаємо бібліотеку для роботи із файлами
from os import listdir
# Підключаємо бібліотеку для роботи з ексель файлами розширення xlsx, version 2.6.1
from openpyxl import *
from openpyxl import styles
# Підключаємо бібліотеку для роботи з ексель файлами розширення xls, version 1.2.0
from xlrd import *
# Бібліотека для роботи із файлами(форматом) json
import json


# Функція для форматування коду УКТЗЕД у вигляд який відображає бухгалтерська програма 1С
def formate_uktzed(uktaed_float):
    uktzed_clear = str(uktaed_float)[:-2]
    group_first = uktzed_clear[:4]
    group_second = uktzed_clear[4:6]
    group_third = uktzed_clear[6:8]
    group_fourth = uktzed_clear[8:]
    txt = "{} {} {} {}"
    return txt.format(group_first, group_second, group_third, group_fourth)


# Використовується для вияснення необхідності у додавані пустого рядка для кращего вигляду і читаємості
def is_needed_new_line(previous, current):
    if previous is None:
        return False
    return previous != current


border = styles.Border(
    left=styles.Side(border_style='thin', color='FF000000'),
    right=styles.Side(border_style='thin', color='FF000000'),
    top=styles.Side(border_style='thin', color='FF000000'),
    bottom=styles.Side(border_style='thin', color='FF000000')
)

# Створюємо та відкриваємо нову книгу та робочий лист із назвою "New_codes_uktzed" де зберігатимемо потрібні нам дані
wb = Workbook()
ws = wb.create_sheet("New_codes_uktzed", 0)
New_codes_uktzed = wb["New_codes_uktzed"]
# Виключаємо відображення сітки в документі
New_codes_uktzed.sheet_view.showGridLines = False
# Створюємо перший рядок з назвами колонок
New_codes_uktzed.cell(row=1, column=1, value="Код товара")
New_codes_uktzed.cell(row=1, column=2, value="Артикул")
New_codes_uktzed.cell(row=1, column=3, value="Назва товара згідно інвойсу")
New_codes_uktzed.cell(row=1, column=4, value="Код товара УКТЗЕД")
New_codes_uktzed.cell(row=1, column=5, value="Технічна довідка")
# Встановлюємо ширину кожної колонки
New_codes_uktzed.column_dimensions['A'].width = 11
New_codes_uktzed.column_dimensions['B'].width = 30
New_codes_uktzed.column_dimensions['C'].width = 27
New_codes_uktzed.column_dimensions['D'].width = 19
New_codes_uktzed.column_dimensions['E'].width = 100

# Дізнаємося з якими файлами будемо працювати та записуємо їх у список
allFiles = listdir(".")
allXlsFilesInCurrentFolder = []  # Список всіх файлів xls з новими кодами УКТЗЕД
for item in allFiles:
    pos = item.find(".")
    filenameExtension = item[(pos + 1):]
    if filenameExtension == "xls":
        allXlsFilesInCurrentFolder.append(item)

# Створюємо обєкт для запису в нього всіх кодів товара та відповідних кодів УКТЗЕД
code_codeUKTZED = {}

# Відкриваємо кожен файл із списку allXlsFilesInCurrentFolder, зчитуємо дані та записуємо у новий файл
numberOfRowRead = 5  # Номер рядку з якого починаємо записувати дані
numberOfRowWrite = 2  # Номер рядку з якого починаємо записувати дані
for item in allXlsFilesInCurrentFolder:
    # Відкриваємо перший файл
    temp = open_workbook(item)
    sh = temp.sheet_by_index(0)
    # Змінна яка встановлює з якого рядка починаємо зчитувати інформацію, значення за замовченням = 5
    start_row = 5
    print(sh.nrows)
    # Знаходимо рядок з якого починати зчитувати інформацію
    for i in range(0, sh.nrows):
        val = sh.cell_value(i, 0)
        reference_value = 'код товара'
        if val == reference_value:
            start_row = i + 1
            break

    previous_uktzed = None
    # Циклом проходився по всіх рядках відкритого файла
    for doc in range(start_row, sh.nrows):
        # Код УКТЗЕД
        uktzed = formate_uktzed(sh.cell_value(doc, 10))
        # Додаткова перевірка на необхідність додавання нового рядку для кращої читаємості
        if is_needed_new_line(previous_uktzed, uktzed):
            numberOfRowWrite = numberOfRowWrite + 2
        previous_uktzed = uktzed
        New_codes_uktzed.cell(numberOfRowWrite, column=4, value=uktzed).border = border
        # Код товара
        code = sh.cell_value(doc, 0)
        New_codes_uktzed.cell(numberOfRowWrite, column=1, value=code).border = border
        # Артикул
        art = sh.cell_value(doc, 1)
        New_codes_uktzed.cell(numberOfRowWrite, column=2, value=art).border = border
        # Назва товара згідно інвойсу
        name = sh.cell_value(doc, 2)
        New_codes_uktzed.cell(numberOfRowWrite, column=3, value=name).border = border
        # Записуємо код товара і код УКТЗЕД в обєкт "code_codeUKTZED" для подальшого зберігання у файл ".py"
        code_codeUKTZED[code] = uktzed
        # Технічна довідка
        technical_descriptions = sh.cell_value(doc, 7)
        New_codes_uktzed.cell(numberOfRowWrite, column=5, value=technical_descriptions).border = border
        # Збільшуємо лічильник рядків для записуємого файла
        numberOfRowWrite = numberOfRowWrite + 1

# Зберігаємо ексель книгу
wb.save('All_new_codes_uktzed.xlsx')

# Видаляємо пусті ключі із обєкта
code_codeUKTZED.pop('')

# Записуємо данні групи товара у файл
f = open('All_new_codes_uktzed.json', 'w')
json_string = json.dumps(code_codeUKTZED, indent=4)
f.write(json_string)
f.close()
